"""Michelle's occupancy is shown beside ours, never instead of it.

WHY BOTH, AND WHY NO VERDICT

Scorecard Pro computes occupancy from the P&L: physical as
1 - |vacancy loss| / GPR, economic as NRI / GPR. Both are dollar-weighted.
Michelle's template states its own figures, almost certainly unit-based
(occupied units / total units). They disagree on the one month where an
Eagle Rock sheet and its P&L overlap -- hers 0.6044 / 0.5419, ours
0.5687 / 0.4429 -- and neither is wrong, because they are not the same
measurement. So the page carries both with provenance and computes no
variance. The disagreement is the information.

THREE THINGS THAT COULD HAVE GONE WRONG SILENTLY, EACH PINNED HERE

1. MONTH ALIGNMENT. The T12 KPIs sheet is a snapshot, not something
   regenerated per upload, and its headers are TEXT so they never coerce
   to dates. Measured on the real files: Jackson's sheet covers 5/24-12/24
   against a P&L of Aug 2025 - Jul 2026 -- ZERO overlap -- while Eagle
   Rock overlaps on four months. Pairing by position would have lined one
   period up against another and looked entirely plausible.

2. THE EXISTING MONTH NORMALISER IS WRONG FOR THESE HEADERS.
   PnLParser.normalize_month finds the month and then takes the month's
   own digits as the year: '10/24' -> 'Oct 2010', '11/24' -> 'Nov 2011',
   '12/24' -> 'Dec 2012', '06/25' -> 'Jun 2006'. Eagle Rock's sheet begins
   at '10/24'. Reusing it would have misfiled three months by a decade and
   then reported no overlap, which looks exactly like the Jackson case.

3. ROWS FOUND BY LABEL, NOT POSITION. Both workbooks happen to put
   'Physical occupancy' at A2 and 'Economic Occupancy' at A3. That holds
   until somebody inserts a row.
"""

import unittest
from pathlib import Path

import openpyxl

from tools.scorecard_pro.parsing import (
    PnLParser,
    ScorecardKpiParser,
    align_stated_occupancy,
    kpi_month_key,
)

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "tools" / "scorecard_pro.html"


def workbook(tmpdir, rows, headers=("5/24", "6/24", "7/24"), sheet="T12 KPIs"):
    """A minimal T12 KPIs sheet. `rows` is a list of (label, [values])."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.cell(1, 1, "KPI")
    for i, h in enumerate(headers, start=2):
        ws.cell(1, i, h).number_format = "@"
    for r, (label, values) in enumerate(rows, start=2):
        ws.cell(r, 1, label)
        for i, v in enumerate(values, start=2):
            ws.cell(r, i, v)
    path = Path(tmpdir) / "sc.xlsx"
    wb.save(path)
    return path


class MonthHeadersParseCorrectlyTests(unittest.TestCase):
    """The dedicated normaliser exists because the shared one is wrong."""

    def test_single_digit_months(self):
        self.assertEqual(kpi_month_key("5/24"), "May 2024")
        self.assertEqual(kpi_month_key("1/25"), "Jan 2025")

    def test_the_cases_the_shared_normaliser_gets_wrong(self):
        for raw, expected in (("10/24", "Oct 2024"), ("11/24", "Nov 2024"),
                              ("12/24", "Dec 2024"), ("06/25", "Jun 2025")):
            with self.subTest(raw=raw):
                self.assertEqual(kpi_month_key(raw), expected)

    def test_the_shared_normaliser_really_is_wrong_here(self):
        """Pinned so the claim above is checked, not asserted.

        If this ever starts passing, PnLParser.normalize_month has been
        fixed and the dedicated one may be reconsidered.
        """
        shared = PnLParser.__new__(PnLParser)
        self.assertEqual(shared.normalize_month("10/24"), "Oct 2010")
        self.assertNotEqual(shared.normalize_month("10/24"), "Oct 2024")

    def test_four_digit_years_and_junk(self):
        self.assertEqual(kpi_month_key("1/2025"), "Jan 2025")
        for bad in ("junk", "", None, "13/24", "0/24"):
            self.assertIsNone(kpi_month_key(bad))


class RowsAreFoundByLabelTests(unittest.TestCase):
    def setUp(self):
        import tempfile
        self.tmp = tempfile.mkdtemp()

    def test_it_reads_both_rows(self):
        path = workbook(self.tmp, [("Physical occupancy", [0.5, 0.6, 0.7]),
                                   ("Economic Occupancy", [0.4, 0.5, 0.6])])
        p = ScorecardKpiParser(path)
        p.parse()
        self.assertEqual(p.get_data()["physical"]["May 2024"], 0.5)
        self.assertEqual(p.get_data()["economic"]["Jul 2024"], 0.6)
        self.assertEqual(sorted(p.get_diagnostics()["rows_found"]),
                         ["economic", "physical"])

    def test_an_inserted_row_does_not_break_it(self):
        """Position matching would read the wrong row here."""
        path = workbook(self.tmp, [("Something New", [9, 9, 9]),
                                   ("Physical occupancy", [0.5, 0.6, 0.7]),
                                   ("Economic Occupancy", [0.4, 0.5, 0.6])])
        p = ScorecardKpiParser(path)
        p.parse()
        self.assertEqual(p.get_data()["physical"]["May 2024"], 0.5)
        self.assertEqual(p.get_data()["economic"]["May 2024"], 0.4)

    def test_a_missing_sheet_is_reported_not_raised(self):
        path = workbook(self.tmp, [("Physical occupancy", [0.5])],
                        headers=("5/24",), sheet="Something Else")
        p = ScorecardKpiParser(path)
        p.parse()
        self.assertFalse(p.get_diagnostics()["sheet_found"])
        self.assertEqual(p.get_data()["physical"], {})

    def test_a_missing_row_is_reported(self):
        path = workbook(self.tmp, [("Physical occupancy", [0.5])],
                        headers=("5/24",))
        p = ScorecardKpiParser(path)
        p.parse()
        self.assertIn("economic", p.get_diagnostics()["rows_missing"])

    def test_non_numeric_cells_are_skipped(self):
        """Jackson's first physical cell is the string 'N/A'."""
        path = workbook(self.tmp, [("Physical occupancy", ["N/A", 0.6, 0.7])])
        p = ScorecardKpiParser(path)
        p.parse()
        self.assertNotIn("May 2024", p.get_data()["physical"])
        self.assertEqual(p.get_data()["physical"]["Jun 2024"], 0.6)


class AlignmentIsByMonthTests(unittest.TestCase):
    STATED = {"physical": {"Jun 2025": 0.6044, "Jul 2025": 0.674},
              "economic": {"Jun 2025": 0.5419, "Jul 2025": 0.6036}}

    def test_overlapping_months_are_paired(self):
        out = align_stated_occupancy(self.STATED, ["Jun 2025", "Jul 2025", "Aug 2025"])
        self.assertEqual(out["overlap_count"], 2)
        self.assertEqual(out["physical"]["Jun 2025"], 0.6044)

    def test_no_overlap_yields_nothing_rather_than_something(self):
        """Jackson's real case: her sheet is a different year entirely."""
        out = align_stated_occupancy(self.STATED, ["Aug 2025", "Sep 2025"])
        self.assertEqual(out["overlap_count"], 0)
        self.assertEqual(out["months"], [])

    def test_it_never_borrows_a_figure_from_another_month(self):
        out = align_stated_occupancy(self.STATED, ["Aug 2025"])
        self.assertNotIn("Aug 2025", out["physical"])


class StaleSourceIsDetectedTests(unittest.TestCase):
    """Jackson's economic figures are IMPORTRANGE caches Excel cannot refresh."""

    def setUp(self):
        import tempfile
        self.tmp = tempfile.mkdtemp()

    def test_a_plain_workbook_is_not_flagged(self):
        path = workbook(self.tmp, [("Physical occupancy", [0.5]),
                                   ("Economic Occupancy", [0.4])],
                        headers=("5/24",))
        p = ScorecardKpiParser(path)
        p.parse()
        self.assertFalse(p.get_diagnostics()["stale_source"])

    def test_an_importrange_formula_is_flagged(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "T12 KPIs"
        ws.cell(1, 1, "KPI")
        ws.cell(1, 2, "5/24").number_format = "@"
        ws.cell(2, 1, "Physical occupancy")
        ws.cell(2, 2, 0.5)
        ws.cell(3, 1, "Economic Occupancy")
        ws.cell(3, 2, '=IFERROR(IMPORTRANGE("https://x","Scorecard!m23"),0.8185)')
        path = Path(self.tmp) / "imp.xlsx"
        wb.save(path)
        p = ScorecardKpiParser(path)
        p.parse()
        self.assertTrue(p.get_diagnostics()["stale_source"])


class TheCardShowsBothAndPicksNeitherTests(unittest.TestCase):
    def setUp(self):
        text = TEMPLATE.read_text(encoding="utf-8")
        start = text.index("function renderStatedOccupancy")
        block = text[start:start + 4200]
        # Strip JS comments before asserting on wording -- the comments
        # here explain the very phrasing under test, and a comment quoting
        # a string is not the page saying it. Second instance of that
        # collision; see HANDOFF.
        self.block = "\n".join(
            line for line in block.splitlines()
            if not line.strip().startswith("//"))
        self.full = text

    def test_both_sources_are_labelled(self):
        self.assertIn("your workbook", self.block)
        self.assertIn("this P&amp;L", self.block)

    def test_no_variance_is_computed(self):
        """No arithmetic between the two figures, and no verdict.

        Banning the word "correct" outright was the first attempt and it
        was too crude: the page says "neither is treated as the correct
        one", which is the refusal to pick, not a verdict. The check is
        for a computed difference and for language that ranks them.
        """
        low = self.block.lower()
        for verdict in ("variance", "discrepanc", "more accurate", "is wrong"):
            self.assertNotIn(verdict, low)
        # No subtraction between hers and ours anywhere in the renderer.
        self.assertNotIn("stated.physical[m] -", self.block)
        self.assertNotIn("- stated.physical", self.block)
        self.assertNotIn("our.occupancy -", low)

    def test_it_says_in_words_that_neither_is_the_answer(self):
        self.assertIn("neither is treated as the correct one", self.block)

    def test_the_no_overlap_case_says_so_explicitly(self):
        self.assertIn("No months in common", self.block)
        self.assertIn("do not describe", self.block)

    def test_a_missing_computed_figure_is_named_not_blank(self):
        """A blank beside percentages reads as zero."""
        self.assertIn("not computed", self.block)

    def test_the_stale_cache_is_disclosed(self):
        self.assertIn("stale_source", self.block)
        self.assertIn("cannot refresh", self.block)

    def test_the_card_is_wired_into_render(self):
        self.assertIn("renderStatedOccupancy(analysis)", self.full)
        self.assertIn('id="stated-occupancy-card"', self.full)


if __name__ == "__main__":
    unittest.main()
