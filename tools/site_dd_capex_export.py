"""
FIRE Capital Tools - Site DD capex export.

Turns an assessment's findings into a capital budget: PDF via matplotlib
PdfPages and XLSX via openpyxl, the same two mechanisms every other
export here uses.

THE PROVENANCE COLUMN IS THE POINT

Every line says where its number came from:

    Inspector estimate    somebody stood in the room and judged it
    Researched average    a national figure from tools/site_dd_reference_costs
    No estimate           nothing priced it, and the line says so

Those are three different kinds of claim and a budget that rendered them
identically would let a national average acquire the authority of a
site visit. The three are separated in the totals as well as the rows,
because "of this $84,000, $61,000 is national averages nobody has
checked" is the sentence that decides whether the number is usable.

WHAT IS NOT INCLUDED

A finding with no cost contributes nothing to the total, but IS still
listed with "no estimate". Dropping it would make the budget look
complete when it is not, which is the failure mode that matters: an
underlying repair that never got priced is invisible, and the total
reads as the whole job.
"""

from __future__ import annotations

import textwrap
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt                       # noqa: E402
from matplotlib.backends.backend_pdf import PdfPages   # noqa: E402

from tools import site_dd_checklist as cl              # noqa: E402
from tools import site_dd_conditions as cond           # noqa: E402
from tools import site_dd_costs as costs               # noqa: E402
from tools import site_dd_reference_costs as refcosts  # noqa: E402
from tools import underwriting_capex as ucx            # noqa: E402

PAGE_SIZE = (11, 8.5)
INK = "#1a2744"
MUTED = "#6b7280"
BODY = "#111827"
WARN = "#b45309"

# PDF layout. Rows are no longer a uniform height -- a line with no total
# carries its explanation underneath -- so pages are filled by measured
# height rather than by a fixed row count. ROWS_PER_PAGE is gone with it:
# it could not describe a page whose rows differ in size, and keeping it as
# a second opinion about how much fits would just be a number to disagree
# with _paginate().
SUMMARY_TOP = 0.865      # where the summary block starts on page 1
ROW_H = 0.026            # one table row
NOTE_H = 0.0185          # one wrapped line of a per-row explanation
HEAD_GAP = 0.024         # column headings to first row
FOOTER_TOP = 0.085       # nothing is drawn below this; footer sits at 0.05
NOTE_WRAP = 120          # characters per note line at fontsize 7.5

SOURCE_COLUMN = {
    costs.SOURCE_MANUAL: "Inspector estimate",
    costs.SOURCE_REFERENCE: "Researched average",
    costs.SOURCE_NONE: "No estimate",
}

# THIS BUCKET NAMES A PERMANENT STATE, NOT A PENDING ONE
#
# It used to read "Researched rate, not yet measured", which described a
# to-do: go and measure, then this line will total. Nobody will. Michelle
# was explicit -- "DON'T WORRY ABOUT CALCULATING PAINT. WE JUST NEED TO
# DETERMINE THE CONDITIONS OF THE WALLS/CEILINGS" -- so the walk records
# condition and the quantity stays NULL forever.
#
# That is not a matter of habit, it is structural: site_dd.py:898 stores
# `quantity` only for a KIND_NUMBER item, and the three KIND_NUMBER items
# are readings (gallons, years), none of them priced by a rate. No route
# anywhere writes a measured area. A rate-priced line therefore CANNOT
# acquire a quantity through this app, and a bucket name promising it will
# is a message describing an action nobody can take.
#
# Defined here, once, because the PDF and the XLSX both print it and a
# budget that named this bucket two different ways would be two different
# claims about the same money.
BUCKET_PRICED_BY_SCOPE = "Priced by scope, not by this walk"
BUCKET_NO_FIGURE = "No researched figure"


def _line_label(item_key, instance_label, labels, scope_label=None):
    """What the budget calls this line.

    "ROOF COVERING — BUILDING 3", NOT "BUILDING 3"

    The instance label used to REPLACE the item's name, which is right for
    a custom item -- somebody adds "Gazebo" and the line should say Gazebo
    -- and wrong for a catalogue item. Michelle asked to record which
    buildings the problems are on: "it varies by property, but typically
    buildings have numbers associated with them." A budget line reading
    "Building 3", with no indication that the $35,000 is a roof, is a line
    nobody can price.

    So the two are joined when both exist and differ, and either one alone
    is used when it is all there is. A custom item has no catalogue label
    to join, so it still reads as just its own name.

    AND THE SCOPE OF WORK, IN PARENTHESES, BECAUSE THREE PARTS NEED TWO
    DIFFERENT SEPARATORS

    A line can now carry three facts: what the item is, WHERE it is, and
    WHICH JOB it needs. `Toilet — Powder room (Replace seat)`.

    The design proposed suffixing the scope after an em dash, which was
    written when the instance label REPLACED the item name. It no longer
    does -- it is joined with an em dash -- so a suffix would produce
    `Toilet — Powder room — replace seat`: two identical separators, and
    no way to tell the place from the job. Parentheses for the scope keep
    every shipped label byte-identical (`Roof covering — Building 3` is
    unchanged and its tests still pass) while giving the third fact a
    marker of its own.

    The alternative -- parenthesise the PLACE and em-dash the job, which
    is what the design sketch would become if followed literally -- reads
    at least as well and was rejected only because it would rewrite labels
    that are already shipped and verified against production data. Nothing
    about the punctuation is load-bearing; not changing working output is.

    `scope_label` is the inspector's own words, already looked up, and is
    None when the scope must not be shown -- see the call site, where the
    rule is that it never repeats what the `state` column is saying.
    """
    known = (labels or {}).get(item_key)
    instance = (instance_label or "").strip()
    scope = (scope_label or "").strip()
    if not instance:
        base = known or item_key
    elif not known or known.strip().lower() == instance.lower():
        base = instance
    else:
        base = f"{known} — {instance}"
    return f"{base} ({scope})" if scope else base


def build_lines(findings: list[dict[str, Any]], labels: dict[str, str] | None = None,
                detail_labels: dict[tuple[str, str], str] | None = None
                ) -> list[dict[str, Any]]:
    """Budget rows, with quantity as the instance count.

    Forty toilets are one line of quantity 40, not forty lines of one.
    This is the grouping site_dd_costs.to_capex_lines() has always
    implemented for the Underwriting hand-off; the export was written
    separately and hard-coded quantity to 1, so a unit cost entered by
    hand produced a line total of exactly that unit cost however many of
    the thing there were. The two paths now agree.

    WHAT IS IN THE GROUPING KEY, AND WHY IT IS MORE THAN (area, room, item)

    to_capex_lines() groups on (area, room, item) alone and takes the
    first non-null cost it finds. That is safe when every instance is
    priced the same and silently wrong when they are not: two toilets at
    $450 and $600 would become "Toilet x2" at whichever price came first,
    and $300 would leave the budget without a trace.

    So condition, detail, unit cost and provenance join the key. Instances
    that are genuinely the same collapse into one line with a quantity;
    instances that differ in what is wrong with them or what they cost
    stay visible as separate lines. Nothing can be absorbed into a
    quantity unless it is interchangeable with the rows beside it.

    The total is not computed here. It comes from
    underwriting_capex.line_total(), which is the function that already
    owns quantity x unit cost for the whole app -- writing a second
    multiplication here would create two numbers that can disagree.
    """
    labels = labels or {}
    detail_labels = detail_labels or {}
    groups: dict[tuple, dict[str, Any]] = {}
    order: list[tuple] = []

    for f in findings or []:
        described = costs.describe(f)
        # DETAIL JOINS THE KEY FOR THE SAME REASON CONDITION DID
        #
        # Since the work-options fix, a choice finding can reach the
        # budget, and for those the detail IS what is wrong: a smoke alarm
        # that is missing and one that needs replacing are the same item,
        # the same room and the same $260, so without this they would
        # collapse into "Smoke alarm x2" and one of the two states would
        # leave no trace. Admitting these findings without widening the
        # key would have replaced a silent drop with a silent merge.
        key = (f.get("area_id"), f.get("room_id"), f.get("item_key"),
               f.get("condition"), f.get("detail"),
               described["cost"], described["source"],
               (f.get("instance_label") or "").strip())
        if key not in groups:
            groups[key] = {"rows": [], "first": f, "described": described}
            order.append(key)
        groups[key]["rows"].append(f)

    out = []
    for key in order:
        group = groups[key]
        f, described = group["first"], group["described"]
        cat = costs.capex_category(f)

        # WHICH KIND OF QUANTITY THIS LINE TAKES
        #
        # A per-item figure is multiplied by the instance count, which is
        # what the grouping above produced and is unchanged.
        #
        # A rate -- dollars per square foot, per linear foot -- cannot be.
        # Multiplying $5.75/sqft by "one wall" yields $5.75, which is the
        # rate wearing a total's clothing. Rates are therefore priced only
        # from a measured quantity recorded against the finding, and when
        # there is none the line carries its rate, no total, and a reason
        # naming the measurement it needs.
        # THE UNIT BELONGS TO THE ITEM, NOT TO WHO PRICED IT
        #
        # This used to read the unit off the _reference object, which is
        # attached only when a researched cost was APPLIED. An inspector
        # typing their own figure on walls_ceiling therefore produced no
        # _reference, fell through to "each", and got $5.75 multiplied by
        # an instance count -- the original bug, arriving through the
        # manual door instead of the researched one.
        #
        # walls_ceiling is a per-square-foot item whoever prices it. So
        # the unit is looked up from the item, and a manual figure on a
        # rate item is a rate.
        ref = f.get("_reference")
        unit = getattr(ref, "unit", None)
        if unit is None:
            # The finding's own detail, unconditionally.
            #
            # This used to read `f.get("detail") if item_key == "flooring"
            # else None`, which passed None on every call it was written
            # for -- a `flooring` finding is KIND_CONDITION and its detail
            # is always NULL, because the material lives on the sibling
            # `flooring_type` row. It was correct by accident and only
            # because every flooring variant is priced per square foot, so
            # nothing downstream noticed which unit came back.
            #
            # It stops being correct the moment a detail changes the UNIT
            # rather than only the rate, which is exactly what
            # COST_BY_DETAIL invites: "repaint" per square foot and
            # "replace drywall" per job on one item key.
            known = refcosts.for_item(f.get("item_key"), f.get("detail"))
            unit = getattr(known, "unit", None)

        # AND WHEN A PERSON TYPED THE FIGURE, THEY GET TO SAY WHAT IT MEANS
        #
        # Michelle: "seven buildings, two roofs -- I'd want to manually
        # input $35,000 or $50,000 myself." A roof is a per-square-foot
        # item, so her own $35,000 was read as a rate and produced no
        # total: the tool refused to add up the two numbers she had just
        # typed.
        #
        # This does NOT reopen the $5.75 decision. That rule is about the
        # REFERENCE table -- a researched national average is published
        # per square foot and stays a rate whatever anyone selects, which
        # is why this branch tests the source before it looks at the
        # toggle. What changes is the MANUAL case, where the person who
        # typed the number is the only one who knows what it means, and
        # she has already been asked: "yes, please add the toggle for 'per
        # sq ft' or 'per job'. It's worth the extra click to ensure the
        # data is accurate."
        #
        # UNSET STAYS UNSET. A blank toggle falls through untouched and
        # the line behaves exactly as it does today -- rate, no total,
        # reference sentence. Resolving absence to a default is the Part
        # 46 failure and it decides money here.
        if (described["source"] == costs.SOURCE_MANUAL
                and described["cost"] is not None):
            stated = _stated_cost_unit(f)
            if stated is not None:
                unit = stated

        if unit is None and described["cost"] is not None:
            measure = _stated_cost_unit(f)
            if measure is not None:
                unit = measure
            else:
                # A FREEFORM item with a hand-typed cost and no answer to
                # the per-job / per-sq-ft toggle.
                #
                # Nothing in the table describes the item, so there is no
                # unit to inherit, and the toggle -- which exists to ask
                # exactly this -- was left blank. The honest answer is
                # that we do not know what the number means, so the line
                # is not totalled at all.
                #
                # It used to be classified by magnitude: under $15 it was
                # called a rate, over $15 a job price. That guess is what
                # the toggle replaces. Guessing and then silently
                # totalling is the failure this whole line of work exists
                # to remove -- a $5.75 that means "per sq ft" became a
                # $5.75 repaint budget for a whole kitchen.
                unit = None

        instances = float(len(group["rows"]))
        measured = [costs.clean_cost(r.get("quantity")) for r in group["rows"]]
        measured = [m for m in measured if m is not None]
        if refcosts.is_rate(unit):
            quantity = sum(measured) if measured else None
        elif unit is None and described["cost"] is not None:
            # A cost in unstated units. There is no quantity that means
            # anything here: multiplying by the instance count would
            # assume "per job", which is the assumption the toggle exists
            # to stop being made silently.
            quantity = None
        else:
            quantity = instances

        needs = ""
        if described["cost"] is not None and unit is None:
            # Priced by a person, in units nobody stated. The magnitude
            # hint is a HINT: it tells the inspector which answer is
            # likely without letting that likelihood become a total.
            hint = ("It looks like a rate" if refcosts.looks_like_a_rate(described["cost"])
                    else "It looks like a job price")
            needs = (f"Cost entered, unit not specified. ${described['cost']:,.2f} "
                     f"could be a price for the whole job or a rate per square "
                     f"foot, and those differ by orders of magnitude. {hint}, "
                     f"but say so on the finding rather than leaving it to be "
                     f"guessed. Not included in the total until then.")
        elif refcosts.is_rate(unit) and quantity is None:
            # A whole sentence, because this lands in a "Why no estimate"
            # cell that has to be readable on its own: it must say what
            # the figure is, that it is a rate, and why no total follows.
            #
            # IT USED TO PRESCRIBE AN ACTION NOBODY CAN TAKE
            #
            # "Needs a measured floor area in square feet before it can be
            # totalled" is a to-do, and there is no way to do it: no route
            # writes a measured area for a rate-priced item (see
            # BUCKET_PRICED_BY_SCOPE above), and Michelle asked for
            # conditions rather than paint quantities. So the line now
            # states the permanent arrangement -- condition here, rate kept
            # as reference for whoever scopes the bid -- instead of asking
            # for a measurement that will never arrive.
            #
            # The RATE STAYS ON THE LINE. It is real researched data and
            # deleting it would throw away the figure Michelle most wants
            # when she scopes this work. It is labelled reference
            # information and explicitly not a total, which is the whole
            # difference between showing a number and claiming one.
            #
            # The provenance is not flattened. A researched national
            # average and an inspector's own rate are different kinds of
            # claim, and this module exists to keep them apart, so the
            # phrase names which one it is.
            if described["source"] == costs.SOURCE_REFERENCE:
                rate_phrase = (f"Researched reference rate "
                               f"${described['cost']:,.2f} "
                               f"{refcosts.UNIT_LABELS.get(unit, unit)} "
                               f"(national average, "
                               f"{getattr(ref, 'dated', refcosts.RESEARCHED_ON)})")
            else:
                rate_phrase = (f"Inspector's rate ${described['cost']:,.2f} "
                               f"{refcosts.UNIT_LABELS.get(unit, unit)}")
            needs = (f"Priced by scope, not by this walk: the walk records "
                     f"this item's condition, not its area. {rate_phrase}, "
                     f"kept as reference information for scoping the work — "
                     f"not a total, and not included in the budget total.")

        line = {
            "item_key": f.get("item_key"),
            # THE SCOPE IS SHOWN ONLY WHEN `state` IS NOT ALREADY SAYING
            # IT, AND THAT RULE IS DERIVED RATHER THAN LISTED.
            #
            # `state` below prints the CONDITION when there is a valid one
            # and falls back to the DETAIL when there is not. So a choice
            # finding -- a missing alarm -- already has its detail on the
            # row, in that column, and repeating it in the label would
            # give one row saying "Missing" twice.
            #
            # A scope finding is the other case: its condition is valid,
            # so `state` prints "Replace" and the detail is nowhere. That
            # is the one that belongs in the label.
            #
            # Written as a condition on `state`'s own rule rather than as
            # a list of scope items, so adding a seventh scope item needs
            # no edit here and cannot be forgotten.
            "label": _line_label(
                f.get("item_key"), f.get("instance_label"), labels,
                detail_labels.get((f.get("item_key"), f.get("detail")))
                if f.get("detail") and cond.is_valid(f.get("condition"))
                else None),
            "category": cat,
            "category_name": cl.CATEGORY_NAMES.get(cat, "Uncategorised"),
            "condition": f.get("condition"),
            "detail": f.get("detail"),
            # WHAT IS WRONG, IN ONE COLUMN, WHICHEVER KIND OF ITEM IT IS
            #
            # The exports have always printed `condition` here. A choice
            # item has none -- its answer is in `detail` -- so once those
            # findings started reaching the budget the column would have
            # read "—" on every missing alarm and absent appliance: a
            # line asking for $260 without saying what for.
            #
            # The stored value is not shown raw. `not_working` is not a
            # word anybody typed; "Present, not working" is what the form
            # said, and echoing the inspector's own wording back is the
            # difference between reporting and paraphrasing. Falls back to
            # the raw value only when the option is not in the catalogue,
            # which means a stale key rather than a normal reading.
            "state": (cond.label(f.get("condition"))
                      if cond.is_valid(f.get("condition"))
                      else (detail_labels.get((f.get("item_key"),
                                               f.get("detail")))
                            or f.get("detail") or "")),
            "scope": f.get("scope"),
            "unit_cost": described["cost"],
            "unit": unit,
            "unit_label": refcosts.UNIT_LABELS.get(unit, ""),
            "is_rate": refcosts.is_rate(unit),
            "instances": instances,
            "source": described["source"],
            "source_label": SOURCE_COLUMN[described["source"]],
            "quantity": quantity,
            # Left None on purpose: line_total() below derives it. Kept on
            # the row so this line has the same shape as an Underwriting
            # capex line, where an explicit total legitimately overrides
            # quantity x unit cost.
            "total_cost": None,
            "reason": (_unpriced_reason(f.get("item_key"))
                       if described["cost"] is None else needs),
        }
        # None, not 0.0, when it cannot be computed honestly. A zero would
        # sum into the total as though the work were free; None keeps the
        # line visible, keeps its rate on screen, and sends it to the
        # unpriced set where summarize() will report it.
        line["total"] = (ucx.line_total(line)
                         if described["cost"] is not None and quantity is not None
                         else None)
        out.append(line)
    return out


def _stated_cost_unit(finding: dict[str, Any]) -> str | None:
    """The cost unit a person chose on the toggle, or None.

    None means unanswered, and unanswered is a real state: it is not
    resolved to "per job" here or anywhere downstream.

    THE COLUMN IS SHARED AND ONLY SOMETIMES MEANS THIS

    `measure` also carries the unit of a NUMBER item's reading -- a water
    heater's "gal", an HVAC's "yr" -- written by the route from the item
    catalogue rather than by anyone answering a toggle. Today those two
    vocabularies happen not to overlap, but "happen not to" is not a
    guarantee: an item whose reading is an area would put "sqft" in this
    column and silently turn a job price into a rate. So a NUMBER item's
    measure is never read as a cost unit, whatever it says.
    """
    from tools import site_dd_bank as bank
    from tools import site_dd_unit_checklist as uc

    item = bank.every_item().get(finding.get("item_key"))
    if item is not None and item.get("kind") == uc.KIND_NUMBER:
        return None
    measure = (finding.get("measure") or "").strip().lower()
    return measure if measure in refcosts.UNITS else None


def _unpriced_reason(item_key: str | None) -> str:
    """Why this line has no figure -- never the empty string.

    `refcosts.reason()` answers for the items somebody has written a
    reason for, and returns "" for anything else. A freeform item nobody
    has researched therefore reached both exports with an empty "why" cell
    beside an empty total: a line asking to be included in a capital
    budget while saying nothing at all about itself.

    That is the silent-gap shape this module has closed four times over,
    and it was still open here -- found when the PDF started rendering
    reasons and some came back blank.

    The fallback states only what the code has established: no figure was
    recorded on the finding, and the reference table has no entry for the
    item. It does NOT say the work is unpriceable or that no figure
    exists in the world -- neither is known.
    """
    written = refcosts.reason(item_key)
    if written:
        return written
    known = refcosts.for_item(item_key)
    if known is not None:
        # In the table, but no cost reached this line -- so the gap is the
        # finding, not the research.
        return ("No cost was recorded on this finding, so it is listed "
                "but not included in the total.")
    return ("No cost was recorded on this finding and the reference table "
            "has no researched figure for this item, so it is listed but "
            "not included in the total.")


def summarize(lines: list[dict[str, Any]]) -> dict[str, Any]:
    """Totals, split by where the money came from.

    Deliberately three totals rather than one. A single figure would
    average a site visit and a national average into a number that
    describes neither.
    """
    by_source = {k: 0.0 for k in SOURCE_COLUMN}
    by_category: dict[str, float] = {}
    priced: list[dict[str, Any]] = []
    unmeasured: list[dict[str, Any]] = []
    unresearched: list[dict[str, Any]] = []
    for l in lines:
        # THREE BUCKETS, AND THEY MUST NOT COLLAPSE INTO TWO
        #
        # priced                 a figure and a quantity to apply it to
        # researched, unmeasured a real researched rate, nothing measured
        #                        to multiply it by. Needs a tape measure.
        # unresearched           no figure at all. Needs research that may
        #                        not exist in published form.
        #
        # Merging the last two is the same mistake as multiplying a rate
        # by a headcount, one level up: it reports "no cost data" for an
        # item whose cost we know to the cent.
        if l["total"] is not None:
            priced.append(l)
            by_source[l["source"]] += l["total"]
            if l["total"]:
                by_category[l["category_name"]] = (
                    by_category.get(l["category_name"], 0.0) + l["total"])
        elif l.get("is_rate") and l["unit_cost"] is not None:
            unmeasured.append(l)
        else:
            unresearched.append(l)

    priced_total = sum(by_source.values())
    # None, not 0.0, when there are lines and none of them could be
    # priced. "$0.00" beside an unmeasured line reads as "this costs
    # nothing" rather than "this cost is not known yet" -- the line
    # refuses to state a number and the summary must not state one on its
    # behalf.
    #
    # An EMPTY budget is different and really is zero: nothing was
    # recorded as needing work, which is a finding, not a gap.
    total = priced_total if (priced or not lines) else None
    unpriced = unmeasured + unresearched
    return {
        "total": total,
        "priced_total": priced_total,
        # True whenever the total describes only part of the work, so a
        # caller cannot render it as a finished budget by accident.
        "total_is_partial": bool(unpriced),
        "by_source": by_source,
        "by_category": dict(sorted(by_category.items(),
                                   key=lambda kv: -kv[1])),
        "priced": priced,
        "priced_count": len(priced),
        "unpriced": unpriced,
        "unpriced_count": len(unpriced),
        "unmeasured": unmeasured,
        "unmeasured_count": len(unmeasured),
        "unresearched": unresearched,
        "unresearched_count": len(unresearched),
        "line_count": len(lines),
        # Share OF THE PRICED TOTAL, and None when there is no priced
        # total. Zero would claim we hold no research, which is false when
        # every unpriced line is a researched rate waiting on a
        # measurement.
        "researched_pct": ((by_source[costs.SOURCE_REFERENCE] / priced_total * 100)
                           if priced_total else None),
        "researched_on": refcosts.RESEARCHED_ON,
        "coverage_sentence": coverage_sentence(
            len(priced), len(lines), len(unmeasured), len(unresearched)),
    }


def coverage_sentence(priced: int, total_lines: int, unmeasured: int,
                      unresearched: int) -> str:
    """One sentence naming all three buckets, in Michelle's words.

    Written once, here, because the PDF and the XLSX must not be able to
    describe the same budget differently -- and because the sentence is
    the part that stops a partial total being read as a finished one.
    """
    def lines_(n):
        return f"{n} line" if n == 1 else f"{n} lines"

    if not total_lines:
        return "No items were recorded as needing work."
    if priced == total_lines:
        return (f"All {lines_(total_lines)} priced. This total is the whole "
                f"recorded budget.")
    bits = []
    if unmeasured:
        # "a researched rate but nothing measured to apply it to" described
        # a gap waiting to be filled. It is not waiting: no route records a
        # measured area, so this is the settled state of these lines and the
        # sentence says so. The rate is still shown -- it is reference
        # information for scoping, not a total.
        bits.append(f"{lines_(unmeasured)} "
                    f"{'is' if unmeasured == 1 else 'are'} priced by scope "
                    f"rather than by this walk, with the rate shown for "
                    f"reference and not totalled")
    if unresearched:
        bits.append(f"{lines_(unresearched)} "
                    f"{'has' if unresearched == 1 else 'have'} no researched "
                    f"figure at all")
    if not priced:
        # Not "can be priced YET". Where every unpriced line is priced by
        # scope, there is nothing pending -- and where a line is genuinely
        # unresearched there is, so the sentence states the fact common to
        # both rather than a schedule that is only true of one.
        return ("Nothing here could be totalled, so there is NO total: "
                + " and ".join(bits) + ".")
    return (f"This total covers {priced} of {lines_(total_lines)} and is NOT "
            f"the full budget: " + " and ".join(bits) + ".")


def _money(value: Any) -> str:
    return "—" if value in (None, "") else f"${float(value):,.0f}"


def _qty(value: Any) -> str:
    """Whole counts read as counts: 40, not 40.0."""
    if value in (None, ""):
        return "—"
    number = float(value)
    return f"{number:,.0f}" if number == int(number) else f"{number:,.2f}"


def _summary_ops(summary: dict[str, Any]) -> list[tuple[float, str, dict, float]]:
    """The "where these numbers come from" block, as positioned draw ops.

    MEASURED AND DRAWN FROM ONE LIST, ON PURPOSE

    The table below it starts wherever this block ends, and pagination now
    needs that height BEFORE anything is drawn. Computing the height in one
    place and drawing in another is how two numbers that must agree stop
    agreeing -- the same failure the bucket-name constant was created to
    prevent, one layer up. So this returns the ops and the caller both
    measures and renders them.

    Each op is (x, text, text_kwargs, dy_consumed_after).
    """
    ops: list[tuple[float, str, dict, float]] = []
    ops.append((0.06, "Where these numbers come from",
                {"fontsize": 10, "fontweight": "bold", "color": INK}, 0.028))

    have_total = summary["total"] is not None
    # Three buckets, three rows. The last two used to share a line reading
    # "N item(s), not costed", which said the same thing about an item
    # priced at $5.75/sqft and an item nobody has ever researched.
    rows = [
        (SOURCE_COLUMN[costs.SOURCE_MANUAL],
         _money(summary["by_source"][costs.SOURCE_MANUAL]) if have_total else "—",
         BODY),
        (SOURCE_COLUMN[costs.SOURCE_REFERENCE],
         _money(summary["by_source"][costs.SOURCE_REFERENCE]) if have_total else "—",
         BODY),
        (BUCKET_PRICED_BY_SCOPE, f"{summary['unmeasured_count']} item(s)", WARN),
        (BUCKET_NO_FIGURE, f"{summary['unresearched_count']} item(s)", MUTED),
    ]
    for name, value, colour in rows:
        ops.append((0.06, name, {"fontsize": 9, "color": BODY}, 0.0))
        ops.append((0.42, value, {"fontsize": 9, "color": colour}, 0.024))

    ops.append((0.06, "", {}, 0.012))
    for text in textwrap.wrap(summary["coverage_sentence"], 108):
        ops.append((0.06, text, {"fontsize": 8.5, "color": WARN}, 0.020))

    if summary["researched_pct"] is not None:
        tail = (f"{summary['researched_pct']:.0f}% of the priced subtotal is "
                f"researched national averages ({summary['researched_on']}), "
                f"not quotes.")
    else:
        tail = (f"Researched national averages ({summary['researched_on']}) are "
                f"used where a figure exists; none could be applied here.")
    ops.append((0.06, tail, {"fontsize": 8.5, "color": MUTED}, 0.020))
    return ops


def _note_lines(row: dict[str, Any]) -> list[str]:
    """The per-line explanation, wrapped, or nothing.

    WHY A LINE WITH NO TOTAL EXPLAINS ITSELF ON THE PAGE

    The PDF printed an em dash in the Total column and left it at that.
    `reason` was rendered only by build_xlsx, so someone reading the budget
    on paper saw "Walls & ceiling ... $5.75 ... per sq ft ... —" with no
    statement of why, while the same budget in Excel carried a full
    sentence. The summary paragraph at the top named the bucket, but a
    reader working down a table does not carry a paragraph three inches up
    with them, and on page two it is not even present.

    An empty Total with no explanation beside it is exactly the silent-gap
    shape this module has closed four times.

    The string is `row["reason"]` VERBATIM -- the same value build_xlsx
    writes into its "Why no estimate" column. It is not re-worded for the
    PDF, because two documents describing one budget in two different
    sentences is the divergence this file keeps designing against.
    """
    if row["total"] is not None:
        return []
    return textwrap.wrap(row["reason"], NOTE_WRAP) if row["reason"] else []


def _paginate(lines: list[dict[str, Any]], first_top: float,
              later_top: float) -> list[list[dict[str, Any]]]:
    """Flow rows into pages by HEIGHT, because rows are no longer uniform.

    A row carrying a three-line explanation is four times the height of a
    priced one, so the old fixed 26-per-page would have run the last rows
    off the bottom of the page as soon as notes appeared. Nothing warns
    about that: matplotlib draws happily at negative coordinates and the
    text simply is not on the paper.

    A block never splits across pages -- a note belongs on the page its row
    is on, which is the entire point of the change.
    """
    pages: list[list[dict[str, Any]]] = []
    page: list[dict[str, Any]] = []
    top = first_top
    used = 0.0
    for row in lines:
        height = ROW_H + len(_note_lines(row)) * NOTE_H
        available = (top - HEAD_GAP) - FOOTER_TOP
        if page and used + height > available:
            pages.append(page)
            page, used, top = [], 0.0, later_top
            available = (top - HEAD_GAP) - FOOTER_TOP
        page.append(row)
        used += height
    if page or not pages:
        pages.append(page)
    return pages


def build_pdf(path, assessment: dict[str, Any], lines: list[dict[str, Any]],
              summary: dict[str, Any]) -> Path:
    path = Path(path)
    label = assessment.get("property_label") or "Property"

    ops = _summary_ops(summary)
    summary_height = sum(op[3] for op in ops)
    first_top = SUMMARY_TOP - summary_height - 0.035
    pages = _paginate(lines, first_top, SUMMARY_TOP)

    with PdfPages(str(path)) as pdf:
        for page_no, page in enumerate(pages, start=1):
            fig = plt.figure(figsize=PAGE_SIZE)
            fig.text(0.06, 0.94, "Capital Budget", fontsize=16,
                     fontweight="bold", color=INK)
            fig.text(0.06, 0.915, label, fontsize=11, color="#4b5563")
            fig.text(0.94, 0.94,
                     _money(summary["total"]) if summary["total"] is not None
                     else "no priced lines",
                     ha="right", fontsize=15, fontweight="bold",
                     color=INK if summary["total"] is not None else WARN)
            fig.text(0.94, 0.915,
                     (f"{summary['line_count']} items · "
                      f"{summary['priced_count']} priced · "
                      f"{summary['unmeasured_count']} priced by scope · "
                      f"{summary['unresearched_count']} unresearched"),
                     ha="right", fontsize=9, color=MUTED)

            if page_no == 1:
                y = SUMMARY_TOP
                for x, text, kwargs, dy in ops:
                    if text:
                        fig.text(x, y, text, **kwargs)
                    y -= dy
                top = first_top
            else:
                top = SUMMARY_TOP

            # Quantity earns a column now that it can be more than 1: a
            # $600 unit cost beside a $24,000 total is unreadable without
            # the 40 that connects them.
            # "Per" is the unit of measure and it is not decoration. A
            # $5.75 rate beside a blank total only makes sense once the
            # column says "per sq ft"; without it the reader sees a cheap
            # item that failed to add up.
            cols = (0.06, 0.26, 0.42, 0.54, 0.68, 0.76, 0.87, 0.94)
            heads = ("Item", "Category", "Condition", "Source", "Rate",
                     "Per", "Qty", "Total")
            for x, head, align in zip(cols, heads,
                                      ("left",) * 7 + ("right",)):
                fig.text(x, top, head, fontsize=8.5, fontweight="bold",
                         color=MUTED, ha=align)
            y = top - HEAD_GAP

            for row in page:
                fig.text(cols[0], y, textwrap.shorten(str(row["label"]), 30,
                                                      placeholder="…"),
                         fontsize=8.5, color=BODY)
                fig.text(cols[1], y, textwrap.shorten(row["category_name"], 26,
                                                      placeholder="…"),
                         fontsize=8, color=MUTED)
                fig.text(cols[2], y, row["state"] or "—",
                         fontsize=8, color=MUTED)
                fig.text(cols[3], y, row["source_label"], fontsize=8,
                         color=WARN if row["source"] == costs.SOURCE_REFERENCE
                         else (MUTED if row["source"] == costs.SOURCE_NONE else BODY))
                fig.text(cols[4], y, _money(row["unit_cost"]), fontsize=8.5,
                         color=BODY)
                fig.text(cols[5], y, row["unit_label"] or "—", fontsize=8,
                         color=WARN if row["is_rate"] else MUTED)
                qty = row["quantity"]
                fig.text(cols[6], y, _qty(qty), fontsize=8.5,
                         color=BODY if (qty or 0) > 1 else MUTED)
                fig.text(cols[7], y,
                         _money(row["total"]) if row["total"] is not None else "—",
                         fontsize=8.5, color=BODY, ha="right")
                y -= ROW_H

                # The explanation, directly under its own row. Indented so
                # it reads as belonging to the line above rather than as a
                # new line item, and coloured by which bucket it is in --
                # a researched rate held back for want of a measurement is
                # a different situation from an item nobody has priced.
                for note in _note_lines(row):
                    fig.text(cols[0] + 0.012, y, note, fontsize=7.5,
                             color=WARN if (row["is_rate"]
                                            and row["unit_cost"] is not None)
                             else MUTED)
                    y -= NOTE_H

            fig.text(0.06, 0.05,
                     "Researched averages are national figures for budgeting, "
                     "not quotes. Items with no estimate, and rates with no "
                     "measured quantity, are listed but contribute nothing to "
                     "the total.",
                     fontsize=7.5, color=MUTED)
            fig.text(0.94, 0.05, f"Page {page_no} of {len(pages)}",
                     ha="right", fontsize=8, color=MUTED)
            pdf.savefig(fig)
            plt.close(fig)
    return path


def build_xlsx(path, assessment: dict[str, Any], lines: list[dict[str, Any]],
               summary: dict[str, Any],
               labels: dict[str, str] | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Capital Budget"
    bold = Font(bold=True)

    ws.append(["Capital Budget"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([assessment.get("property_label") or ""])
    ws.append([f"Inspected {assessment.get('assessed_on') or '—'}"])
    ws.append([])
    # The label carries the caveat, so the number can never be lifted out
    # of the cell beside it and read as a finished budget.
    if summary["total"] is None:
        ws.append(["Total", "No priced lines — see below"])
    else:
        ws.append(["Priced subtotal" if summary["total_is_partial"] else "Total",
                   summary["total"]])
    for key in (costs.SOURCE_MANUAL, costs.SOURCE_REFERENCE):
        # Dash, not 0.00, when there is no priced subtotal for them to be
        # a part of -- the same reason the total itself declines to be a
        # number here.
        ws.append([SOURCE_COLUMN[key],
                   summary["by_source"][key] if summary["total"] is not None
                   else "—"])
    ws.append([BUCKET_PRICED_BY_SCOPE,
               f"{summary['unmeasured_count']} item(s)"])
    ws.append([BUCKET_NO_FIGURE,
               f"{summary['unresearched_count']} item(s)"])
    ws.append([summary["coverage_sentence"]])
    if summary["researched_pct"] is not None:
        ws.append([f"{summary['researched_pct']:.0f}% of the priced subtotal is "
                   f"researched national averages ({summary['researched_on']}), "
                   f"not quotes."])
    else:
        ws.append([f"Researched national averages ({summary['researched_on']}) "
                   f"are used where a figure exists; none could be applied here."])
    ws.append([])

    header = ["Item", "Category", "Scope", "Condition", "Cost source",
              "Unit cost", "Unit", "Qty", "Total", "Why no estimate"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = bold

    for l in lines:
        ws.append([l["label"], l["category_name"], l["scope"],
                   l["state"], l["source_label"],
                   l["unit_cost"], l["unit_label"], l["quantity"],
                   l["total"], l["reason"]])

    for col, width in zip("ABCDEFGHIJ", (30, 26, 10, 12, 20, 12, 13, 8, 12, 60)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=11, min_col=10, max_col=10):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # Rates that have a researched figure and no measurement to apply it
    # to. Its own sheet rather than a footnote: these are the lines a
    # walk can turn into real money, and each one names what to measure.
    if summary["unmeasured"]:
        nm = wb.create_sheet("Needs measurement")
        nm.append(["Item", "Where", "Rate", "Unit", "What is needed"])
        for cell in nm[1]:
            cell.font = bold
        for l in summary["unmeasured"]:
            nm.append([l["label"], l["scope"], l["unit_cost"],
                       l["unit_label"], l["reason"]])
        for col, width in zip("ABCDE", (30, 12, 12, 14, 60)):
            nm.column_dimensions[col].width = width
        for row in nm.iter_rows(min_row=2, min_col=5, max_col=5):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # The reference table itself, so a reader can audit any figure that
    # appeared above without leaving the file.
    ref = wb.create_sheet("Reference costs")
    ref.append(["Item", "Key", "Unit cost", "Unit", "Sources",
                "How it was derived"])
    for cell in ref[1]:
        cell.font = bold
    for key in sorted(refcosts.REFERENCE_COSTS):
        c = refcosts.REFERENCE_COSTS[key]
        ref.append([(labels or {}).get(c.key, c.key), c.key, c.unit_cost,
                    refcosts.UNIT_LABELS[c.unit],
                    ", ".join(c.sources), c.note])
    for col, width in zip("ABCDEF", (30, 24, 12, 14, 42, 80)):
        ref.column_dimensions[col].width = width
    for row in ref.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # And what has NO figure, with the reason. This sheet is the ask, so
    # it gets the labels a person recognises -- "ADA parking & path of
    # travel", not "ada_parking_path". A list meant to be read by
    # somebody who does not work in this codebase should not be written
    # in its identifiers.
    un = wb.create_sheet("Not priced")
    un.append(["Item", "Key", "Why it has no researched figure"])
    for cell in un[1]:
        cell.font = bold
    for row in refcosts.unpriced_report(labels or {}):
        un.append([row["label"], row["key"], row["reason"]])
    for col, width in zip("ABC", (34, 24, 90)):
        un.column_dimensions[col].width = width
    for r in un.iter_rows(min_row=2, min_col=3, max_col=3):
        r[0].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(str(path))
    return path


def suggested_filename(assessment: dict[str, Any], ext: str) -> str:
    label = "".join(ch if ch.isalnum() or ch in " -_" else ""
                    for ch in (assessment.get("property_label") or "property"))
    label = "-".join(label.split()).lower()[:48] or "property"
    return f"capex-budget-{label}.{ext}"
