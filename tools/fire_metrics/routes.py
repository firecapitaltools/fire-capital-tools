"""FIRE Capital Tools — FIRE Metric search dashboard (Flask routes)."""

from __future__ import annotations

import io
import json
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from flask import Blueprint, abort, current_app, jsonify, render_template, request, send_file
from flask_login import current_user, login_required
from models import User

from fire_metrics.fire_metrics_updater import db as db_module
from fire_metrics.fire_metrics_updater.city_search import find_city_match
from tools import fire_metrics_ai_summary as ai_summary
from tools.fire_metrics.constants import MAX_CRIME_WORKBOOK_BYTES, TOP_CITY_METRICS
from tools.fire_metrics.crime_workbook import (
    _crime_workbook_status,
    _get_crime_workbook_path,
    _validate_crime_workbook_bytes,
)
from tools.fire_metrics.services import (
    _build_fire_score_index,
    _cre_research_model_name,
    _enrich_search_payload_with_fire_score,
    _export_workbook,
    _fetch_top_cities,
    _parse_top_cities_limit,
    _refresh_status,
    _reingest_from_disk,
    _start_refresh,
    _summary_api_key,
    _summary_enabled,
    _summary_model_name,
    _summary_unavailable_response,
)


fire_metrics_bp = Blueprint("fire_metrics", __name__)

ALLOWED_CRE_SELECTION_SOURCES = frozenset({
    "main_city_search",
    "city_analytics",
    "multi_city_search_auto",
})

CITY_ANALYTICS_EXPORT_COLUMNS = [
    ("City", "city", "text"),
    ("State", "state", "text"),
    ("FIRE Score", "fire_score", "float1"),
    ("Data Coverage", "fire_score_coverage", "percent_whole1"),
    ("Population", "population_current", "int"),
    ("Population Growth", "population_growth_recent", "percent2"),
    ("Median Income", "median_income_current", "currency0"),
    ("Income Growth", "median_income_growth_recent", "percent2"),
    ("Home Value", "median_home_value_current", "currency0"),
    ("Home Value Growth", "median_home_value_growth_recent", "percent2"),
    ("Employment", "employment_current", "int"),
    ("Employment Growth", "employment_growth_recent", "percent2"),
    ("Climate Risk", "climate_risk_score", "float1"),
    ("Climate Rating", "climate_risk_rating", "text"),
    ("Crime", "crime_index_score", "float1"),
    ("Crime Rating", "crime_rating", "text"),
    ("Density-Adj. Crime", "density_adjusted_crime_score", "float1"),
    ("Density-Adj. Crime Rating", "density_adjusted_crime_rating", "text"),
    ("Landlord Friendliness", "landlord_friendliness_label", "text"),
]


def _is_admin_user() -> bool:
    if not current_user.is_authenticated:
        return False
    return User.matches_admin_user(current_user.get_id() or "", current_app.config)


def _admin_forbidden_response(*, is_ajax: bool):
    if is_ajax:
        return jsonify({
            "success_message": None,
            "error_message": "Admin access is required for that action.",
        }), 403
    abort(403)


def _safe_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_text(value) -> str:
    return str(value or "").strip()


def _city_label(row: dict) -> str:
    display_name = _safe_text(row.get("display_name"))
    if display_name and "," in display_name:
        return display_name.split(",", 1)[0].strip()
    city = _safe_text(row.get("city"))
    if city and city.lower().endswith(" city"):
        return city[:-5].strip()
    return city


def _normalize_city_analytics_export_rows(payload_rows) -> list[dict]:
    out = []
    if not isinstance(payload_rows, list):
        return out
    for raw in payload_rows:
        if not isinstance(raw, dict):
            continue
        row = {
            "city": _city_label(raw),
            "state": _safe_text(raw.get("state")).upper(),
            "fire_score": _safe_float(raw.get("fire_score")),
            "fire_score_coverage": _safe_float(raw.get("fire_score_coverage")),
            "population_current": _safe_float(raw.get("population_current")),
            "population_growth_recent": _safe_float(raw.get("population_growth_recent")),
            "median_income_current": _safe_float(raw.get("median_income_current")),
            "median_income_growth_recent": _safe_float(raw.get("median_income_growth_recent")),
            "median_home_value_current": _safe_float(raw.get("median_home_value_current")),
            "median_home_value_growth_recent": _safe_float(raw.get("median_home_value_growth_recent")),
            "employment_current": _safe_float(raw.get("employment_current")),
            "employment_growth_recent": _safe_float(raw.get("employment_growth_recent")),
            "climate_risk_score": _safe_float(raw.get("climate_risk_score")),
            "climate_risk_rating": _safe_text(raw.get("climate_risk_rating")),
            "crime_index_score": _safe_float(raw.get("crime_index_score")),
            "crime_rating": _safe_text(raw.get("crime_rating")),
            "density_adjusted_crime_score": _safe_float(raw.get("density_adjusted_crime_score")),
            "density_adjusted_crime_rating": _safe_text(raw.get("density_adjusted_crime_rating")),
            "landlord_friendliness_label": _safe_text(raw.get("landlord_friendliness_label")),
        }
        out.append(row)
    return out


def _city_analytics_export_filename(ext: str) -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"fire_metrics_city_analytics_{stamp}.{ext}"


def _build_city_analytics_xlsx(rows: list[dict]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "City Analytics"

    headers = [col[0] for col in CITY_ANALYTICS_EXPORT_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(col[1]) for col in CITY_ANALYTICS_EXPORT_COLUMNS])

    for idx, (_, _, fmt) in enumerate(CITY_ANALYTICS_EXPORT_COLUMNS, start=1):
        col = ws.column_dimensions[openpyxl.utils.get_column_letter(idx)]
        col.width = 18
        for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
            c = cell[0]
            if c.value is None:
                continue
            if fmt == "currency0":
                c.number_format = "$#,##0"
            elif fmt == "percent2":
                c.number_format = "0.00%"
            elif fmt == "percent_whole1":
                c.number_format = '0.0"%"'
            elif fmt == "float1":
                c.number_format = "0.0"
            elif fmt == "int":
                c.number_format = "#,##0"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _pdf_table_value(row: dict, key: str, fmt: str) -> str:
    value = row.get(key)
    if value is None or value == "":
        return "-"
    if fmt == "currency0":
        return f"${float(value):,.0f}"
    if fmt == "percent2":
        return f"{float(value) * 100:.2f}%"
    if fmt == "percent_whole1":
        return f"{float(value):.1f}%"
    if fmt == "float1":
        return f"{float(value):.1f}"
    if fmt == "int":
        return f"{float(value):,.0f}"
    return _safe_text(value) or "-"


def _build_city_analytics_pdf(rows: list[dict]) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    headers = [col[0] for col in CITY_ANALYTICS_EXPORT_COLUMNS]
    all_rows = [
        [_pdf_table_value(row, key, fmt) for _, key, fmt in CITY_ANALYTICS_EXPORT_COLUMNS]
        for row in rows
    ]
    rows_per_page = 20
    pages = [all_rows[i:i + rows_per_page] for i in range(0, len(all_rows), rows_per_page)] or [[]]

    buffer = io.BytesIO()
    with PdfPages(buffer) as pdf:
        for page_idx, page_rows in enumerate(pages, start=1):
            fig = plt.figure(figsize=(14, 8.5))
            fig.text(0.05, 0.96, "FIRE Metrics - City Analytics", fontsize=14, fontweight="bold", color="#1a2744")
            fig.text(0.95, 0.96, datetime.now(timezone.utc).strftime("Exported %Y-%m-%d"), ha="right", fontsize=9, color="#6b7280")
            fig.text(0.95, 0.03, f"Page {page_idx} of {len(pages)}", ha="right", fontsize=8, color="#6b7280")

            ax = fig.add_axes([0.04, 0.09, 0.92, 0.82])
            ax.axis("off")
            table = ax.table(
                cellText=page_rows,
                colLabels=headers,
                cellLoc="left",
                loc="upper left",
            )
            table.auto_set_font_size(False)
            table.set_fontsize(7)
            table.scale(1, 1.25)
            for (row_i, _), cell in table.get_celld().items():
                if row_i == 0:
                    cell.set_text_props(weight="bold", color="#1a2744")
                    cell.set_facecolor("#eef2fb")
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

    return buffer.getvalue()


def _cre_generation_allowed(payload: dict) -> bool:
    intent = str(payload.get("cre_generation_intent") or "").strip()
    source = str(payload.get("cre_selection_source") or "").strip()
    return intent == "explicit_city_selection" and source in ALLOWED_CRE_SELECTION_SOURCES


def _cre_payload_for_city(
    conn,
    *,
    selected_city: dict,
    data_fingerprint: str,
    model_name: str,
    cre_generation_allowed: bool,
) -> dict:
    """Resolve CRE status for a city, refreshing once when cache is stale and explicitly allowed."""
    try:
        cache_row = db_module.fetch_cached_city_summary(
            conn,
            city=selected_city["city"],
            state=selected_city["state"],
            data_fingerprint=data_fingerprint,
            model_name=model_name,
            prompt_version=ai_summary.PROMPT_VERSION,
        )
    except Exception as exc:
        current_app.logger.warning(
            "FIRE Metrics CRE cache read failed: %s",
            exc.__class__.__name__,
        )
        cache_row = None

    if not cre_generation_allowed:
        return {
            "cre_status": "skipped",
            "cre_summary": "",
            "research_sources": [],
            "cre_failure_category": None,
            "cre_failure_code": None,
            "cre_failure_param": None,
        }

    cached_cre_result_type = ""
    cached_cre_failure_category = None
    cached_cre_failure_code = None
    cached_cre_failure_param = None
    cre_sentences = ""
    research_sources: list[dict] = []
    cre_fresh = False

    if cache_row:
        cached_cre_version = cache_row.get("cre_research_version")
        cre_fresh_base = ai_summary.is_cre_cache_current(
            cache_row.get("cre_generated_at"),
            cached_cre_version,
        )
        cached_cre_result_type = str(cache_row.get("cre_result_type") or "").strip().lower()
        cached_cre_failure_category = str(cache_row.get("cre_failure_category") or "").strip() or None
        cached_cre_failure_code = str(cache_row.get("cre_failure_code") or "").strip() or None
        cached_cre_failure_param = str(cache_row.get("cre_failure_param") or "").strip() or None
        cre_sentences = str(cache_row.get("cre_sentences_text") or "").strip()
        try:
            research_sources = json.loads(cache_row.get("research_sources_json") or "[]") or []
        except (json.JSONDecodeError, ValueError):
            research_sources = []

        if cached_cre_result_type == "success":
            cre_fresh = cre_fresh_base and bool(cre_sentences) and bool(research_sources)
        elif cached_cre_result_type in ("no_data", "failure"):
            cre_fresh = cre_fresh_base
        else:
            cre_fresh = cre_fresh_base and (bool(cre_sentences) or bool(research_sources))

        current_app.logger.info(
            "FIRE CRE cache state: city=%s|%s version=%s expected=%s fresh=%s",
            selected_city["city"], selected_city["state"],
            cached_cre_version, ai_summary.CRE_RESEARCH_VERSION, cre_fresh,
        )

    if cre_fresh:
        cre_status = cached_cre_result_type if cached_cre_result_type in {"success", "no_data", "failure"} else "skipped"
        if cre_status == "skipped" and cre_sentences:
            cre_status = "success"
        return {
            "cre_status": cre_status,
            "cre_summary": cre_sentences if cre_status == "success" else (
                "No relevant research from approved sources." if cre_status == "no_data" else ""
            ),
            "research_sources": research_sources if cre_status == "success" else [],
            "cre_failure_category": cached_cre_failure_category if cre_status == "failure" else None,
            "cre_failure_code": cached_cre_failure_code if cre_status == "failure" else None,
            "cre_failure_param": cached_cre_failure_param if cre_status == "failure" else None,
        }

    api_key = _summary_api_key()
    if not api_key:
        return {
            "cre_status": "failure",
            "cre_summary": "",
            "research_sources": [],
            "cre_failure_category": "config_error",
            "cre_failure_code": None,
            "cre_failure_param": None,
        }

    cre_model = _cre_research_model_name()
    current_app.logger.info(
        "FIRE CRE refresh needed: city=%s state=%s model=%s",
        selected_city["city"], selected_city["state"], cre_model,
    )
    cre_result = ai_summary.openai_cre_research(
        api_key=api_key,
        model_name=cre_model,
        city=selected_city["city"],
        state=selected_city["state"],
        display_name=str(selected_city.get("display_name") or ""),
    )
    result_type = str(cre_result.get("result_type") or "failure").strip().lower() or "failure"
    cached_cre_failure_category = str(cre_result.get("failure_category") or "").strip() or None
    cached_cre_failure_code = str(cre_result.get("failure_code") or "").strip() or None
    cached_cre_failure_param = str(cre_result.get("failure_param") or "").strip() or None

    refreshed_text = cre_result.get("cre_sentences") or ""
    refreshed_sources = cre_result.get("research_sources") or []

    if result_type == "failure":
        import datetime as _dt

        store_at = (
            _dt.datetime.now(_dt.timezone.utc)
            - _dt.timedelta(days=ai_summary.CRE_RESEARCH_TTL_DAYS)
            + _dt.timedelta(minutes=ai_summary.CRE_FAILURE_BACKOFF_MINUTES)
        ).isoformat()
    elif result_type == "no_data":
        import datetime as _dt

        store_at = (
            _dt.datetime.now(_dt.timezone.utc)
            - _dt.timedelta(days=ai_summary.CRE_RESEARCH_TTL_DAYS)
            + _dt.timedelta(hours=ai_summary.CRE_NEGATIVE_CACHE_TTL_HOURS)
        ).isoformat()
    else:
        store_at = cre_result.get("cre_generated_at") or ai_summary.utc_now_iso()

    store_version = cre_result.get("cre_research_version") or ai_summary.CRE_RESEARCH_VERSION

    if cache_row:
        store_cre_text = refreshed_text if result_type != "failure" else (cache_row.get("cre_sentences_text") or "")
        store_sources = refreshed_sources if result_type != "failure" else research_sources
        db_module.update_city_summary_cre_fields(
            conn,
            city=selected_city["city"],
            state=selected_city["state"],
            data_fingerprint=data_fingerprint,
            model_name=model_name,
            prompt_version=ai_summary.PROMPT_VERSION,
            cre_sentences_text=store_cre_text,
            research_sources_json=json.dumps(store_sources),
            cre_generated_at=store_at,
            cre_research_version=store_version,
            cre_result_type=result_type,
            cre_failure_category=cached_cre_failure_category,
            cre_failure_code=cached_cre_failure_code,
            cre_failure_param=cached_cre_failure_param,
        )

    if result_type == "success":
        return {
            "cre_status": "success",
            "cre_summary": refreshed_text,
            "research_sources": refreshed_sources,
            "cre_failure_category": None,
            "cre_failure_code": None,
            "cre_failure_param": None,
        }
    if result_type == "no_data":
        return {
            "cre_status": "no_data",
            "cre_summary": "No relevant research from approved sources.",
            "research_sources": [],
            "cre_failure_category": None,
            "cre_failure_code": None,
            "cre_failure_param": None,
        }
    return {
        "cre_status": "failure",
        "cre_summary": "",
        "research_sources": [],
        "cre_failure_category": cached_cre_failure_category,
        "cre_failure_code": cached_cre_failure_code,
        "cre_failure_param": cached_cre_failure_param,
    }


@fire_metrics_bp.route("/", methods=["GET", "POST"])
@login_required
def index(standalone_mode: bool = False):
    # Computed first (depends only on request headers, can't itself raise)
    # so the outermost except below always knows whether this caller's own
    # JS is going to do res.json() unconditionally on whatever comes back.
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def safe_refresh_status() -> dict:
        # _refresh_status() opens its own SQLite connection -- and, right
        # after _start_refresh() below, so does the refresh subprocess it
        # just spawned (schema-init runs on every get_connection() call).
        # A transient "database is locked" race between those two is
        # possible. If it happens, fall back to a status payload built
        # from what we already know rather than letting the exception
        # propagate out of this view. "running" defaults to False here --
        # if we can't even read the DB, we genuinely don't know, and
        # assuming "not running" lets the user retry rather than looking
        # permanently stuck.
        try:
            return _refresh_status()
        except Exception as exc:
            return {
                "status": "error",
                "running": False,
                "last_refresh_at": None,
                "last_refresh_error": f"Could not read refresh status: {exc}",
                "city_count": 0,
            }

    try:
        status = safe_refresh_status()
        context = {
            "status": status,
            "crime_workbook": _crime_workbook_status(),
            "standalone_mode": bool(standalone_mode),
            "can_manage_fire_metrics_admin": _is_admin_user(),
            "success_message": None,
            "error_message": None,
            "search_query": "",
            "search_payload": None,
            "city_preview": [],
            "google_maps_api_key": current_app.config.get("GOOGLE_MAPS_API_KEY") or "",
            "google_maps_map_id": current_app.config.get("GOOGLE_MAPS_MAP_ID") or "",
        }

        with db_module.get_connection() as conn:
            context["city_preview"] = db_module.fetch_all_cities(conn)[:5]

        if request.method == "GET":
            query = request.args.get("q", "").strip()
            if query:
                context["search_query"] = query
                with db_module.get_connection() as conn:
                    city_index = db_module.build_city_index_payload(conn)
                    excluded_index = db_module.build_excluded_index_payload(conn)
                    score_index = _build_fire_score_index(conn)
                context["search_payload"] = _enrich_search_payload_with_fire_score(
                    find_city_match(query, city_index, excluded_index),
                    score_index,
                )
            return render_template("tools/fire_metrics.html", **context)

        action = request.form.get("action", "")
        admin_actions = {"refresh_all", "refresh_live_only", "rebuild_index"}
        if action in admin_actions and not context["can_manage_fire_metrics_admin"]:
            return _admin_forbidden_response(is_ajax=is_ajax)

        def respond(status_code: int = 200):
            if is_ajax:
                # Admin actions are triggered via fetch() from the page's
                # own JS specifically so they don't navigate away -- a
                # full-page render/redirect here would reload the page and
                # discard whatever search result the user currently has on
                # screen (client-side only state, never persisted
                # server-side).
                payload = {
                    "success_message": context["success_message"],
                    "error_message": context["error_message"],
                    "crime_workbook": _crime_workbook_status(),
                }
                payload.update(safe_refresh_status())
                return jsonify(payload), status_code
            return render_template("tools/fire_metrics.html", **context), status_code

        if action == "refresh_all":
            started = _start_refresh(skip_climate=False, skip_crime=False)
            if started:
                context["success_message"] = "Refresh started in the background. This can take several minutes (climate risk especially, on a cold cache)."
            else:
                context["error_message"] = "A refresh is already running. Check back shortly."
            context["status"] = safe_refresh_status()
            return respond()

        if action == "refresh_live_only":
            # Population/income/home-value/employment only -- skips the slow
            # climate step and the manual/periodic crime step.
            started = _start_refresh(skip_climate=True, skip_crime=True)
            if started:
                context["success_message"] = "Refreshing live metrics (population, income, home value, employment) in the background."
            else:
                context["error_message"] = "A refresh is already running. Check back shortly."
            context["status"] = safe_refresh_status()
            return respond()

        if action == "rebuild_index":
            try:
                results = _reingest_from_disk()
                if not results:
                    context["error_message"] = "No pipeline output files found on disk yet. Run a refresh first."
                else:
                    context["success_message"] = f"Re-ingested from disk: {', '.join(results.keys())}."
            except Exception as exc:
                context["error_message"] = f"Could not re-ingest from disk: {exc}"
            context["status"] = safe_refresh_status()
            return respond()

        context["error_message"] = "Unknown action."
        return respond(status_code=400)
    except Exception as exc:
        # Last-resort guard covering the whole view, including the
        # earliest calls above (status/context/city_preview) that run
        # before respond() even exists yet: if anything in this view
        # raises unexpectedly and the caller is this page's own AJAX JS,
        # it must still get valid JSON back -- otherwise Flask's default
        # HTML error page reaches the browser and res.json() throws a
        # confusing "Unexpected token '<'" instead of showing the real
        # problem. Non-AJAX (plain GET/POST) callers keep the normal Flask
        # error behavior.
        if is_ajax:
            return jsonify({
                "success_message": None,
                "error_message": f"Unexpected error: {exc}",
                "status": "error",
                "running": False,
                "last_refresh_at": None,
                "last_refresh_error": str(exc),
                "city_count": 0,
                "crime_workbook": {"exists": False, "uploaded_at": None},
            }), 500
        raise


@fire_metrics_bp.route("/search")
@login_required
def search():
    query = request.args.get("q", "").strip()
    try:
        with db_module.get_connection() as conn:
            city_index = db_module.build_city_index_payload(conn)
            excluded_index = db_module.build_excluded_index_payload(conn)
            score_index = _build_fire_score_index(conn)
        payload = find_city_match(query, city_index, excluded_index)
        payload = _enrich_search_payload_with_fire_score(payload, score_index)
        payload["query"] = query
        payload["status_meta"] = _refresh_status()
        return jsonify(payload)
    except Exception as exc:
        return jsonify({"status": "error", "query": query, "user_message": f"Search failed: {exc}"}), 500


@fire_metrics_bp.route("/api/top-cities")
@login_required
def top_cities():
    metric = str(request.args.get("metric") or "").strip()
    try:
        limit = _parse_top_cities_limit(request.args.get("limit"))
    except ValueError as exc:
        return jsonify({
            "status": "error",
            "error_code": "invalid_limit",
            "user_message": f"Invalid limit: {exc}",
        }), 400

    if metric not in TOP_CITY_METRICS:
        return jsonify({
            "status": "error",
            "error_code": "invalid_metric",
            "user_message": "Unknown ranking metric.",
        }), 400

    try:
        with db_module.get_connection() as conn:
            spec, cities = _fetch_top_cities(conn, metric_key=metric, limit=limit)
    except Exception as exc:
        current_app.logger.exception("FIRE Metrics top-cities endpoint failed: %s", exc.__class__.__name__)
        return jsonify({
            "status": "error",
            "error_code": "top_cities_failed",
            "user_message": "Top city ranking is currently unavailable.",
        }), 500

    return jsonify({
        "status": "ready",
        "metric": metric,
        "metric_label": spec["label"],
        "direction": spec["direction"],
        "city_count": len(cities),
        "cities": cities,
    })


@fire_metrics_bp.route("/api/city-summary", methods=["POST"])
@login_required
def city_summary():
    payload = request.get_json(silent=True) or {}
    city_key = str(payload.get("city_key") or "").strip()
    city = str(payload.get("city") or "").strip()
    state = str(payload.get("state") or "").strip().upper()

    if not city_key and (not city or not state):
        return jsonify({
            "status": "error",
            "error_code": "invalid_city_identifier",
            "user_message": "City identifier is required.",
        }), 400

    try:
        with db_module.get_connection() as conn:
            selected_city = db_module.fetch_city_by_summary_identity(
                conn,
                city_key=city_key or None,
                city=city or None,
                state=state or None,
            )
            if not selected_city:
                return jsonify({
                    "status": "error",
                    "error_code": "city_not_found",
                    "user_message": "City not found in tracked FIRE Metrics data.",
                }), 404

            all_cities = db_module.fetch_all_included_cities(conn)
            metadata = db_module.get_metadata(conn)

            benchmarks = ai_summary.compute_benchmarks(selected_city, all_cities)

            if not _summary_enabled():
                response_payload = _summary_unavailable_response(
                    selected_city=selected_city,
                    benchmark_data=benchmarks,
                    reason="AI summaries are disabled.",
                    data_refreshed_at=metadata.get("last_refresh_at"),
                )
                response_payload["research_sources"] = []
                response_payload["cre_status"] = "skipped"
                response_payload["cre_failure_category"] = None
                response_payload["cre_failure_code"] = None
                response_payload["cre_failure_param"] = None
                return jsonify(response_payload)

            model_name = _summary_model_name()
            fingerprint_input = ai_summary.fingerprint_payload(
                selected_city=selected_city,
                benchmarks=benchmarks,
                model_name=model_name,
                refresh_last_at=metadata.get("last_refresh_at"),
            )
            data_fingerprint = ai_summary.build_fingerprint(fingerprint_input)

            try:
                cache_row = db_module.fetch_cached_city_summary(
                    conn,
                    city=selected_city["city"],
                    state=selected_city["state"],
                    data_fingerprint=data_fingerprint,
                    model_name=model_name,
                    prompt_version=ai_summary.PROMPT_VERSION,
                )
            except Exception as exc:
                current_app.logger.warning(
                    "FIRE Metrics city-summary cache read failed: %s",
                    exc.__class__.__name__,
                )
                cache_row = None
            if cache_row:
                full_summary = cache_row["summary_text"]

                return jsonify({
                    "status": "ready",
                    "summary": full_summary,
                    "summary_structured": {
                        "strength_sentence": cache_row["strength_sentence"],
                        "weakness_sentence": cache_row["weakness_sentence"],
                        "comparison_sentence": cache_row["comparison_sentence"],
                    },
                    "research_sources": [],
                    "generated_at": cache_row["generated_at"],
                    "data_refreshed_at": metadata.get("last_refresh_at"),
                    "cached": True,
                    "city_key": cache_row["city_key"],
                    "relative_market_profile_score": benchmarks.get("relative_market_profile_score"),
                    "relative_market_profile_percentile": benchmarks.get("relative_market_profile_percentile"),
                    "tracked_city_relative_market_profile_average": benchmarks.get("tracked_city_relative_market_profile_average"),
                    "recommendation_category": benchmarks.get("recommendation_category"),
                    "score": benchmarks.get("selected_overall_score"),
                    "computed_composite_score": benchmarks.get("selected_overall_score"),
                    "tracked_city_average": benchmarks.get("tracked_city_average"),
                    "tracked_city_composite_average": benchmarks.get("tracked_city_average"),
                    "tracked_city_count": benchmarks.get("tracked_city_count"),
                    "percentile": benchmarks.get("selected_percentile"),
                    "source": "cache",
                    "cre_status": "skipped",
                    "cre_failure_category": None,
                    "cre_failure_code": None,
                    "cre_failure_param": None,
                })

            generated_at = ai_summary.utc_now_iso()
            api_key = _summary_api_key()
            if not api_key:
                response_payload = _summary_unavailable_response(
                    selected_city=selected_city,
                    benchmark_data=benchmarks,
                    reason="OPENAI_API_KEY is not configured.",
                    data_refreshed_at=metadata.get("last_refresh_at"),
                )
                response_payload["research_sources"] = []
                response_payload["cre_status"] = "skipped"
                response_payload["cre_failure_category"] = None
                response_payload["cre_failure_code"] = None
                response_payload["cre_failure_param"] = None
                return jsonify(response_payload)

            if not model_name:
                # Keep overview deterministic when summary model is unset,
                # but continue into explicit CRE evaluation below.
                structured = ai_summary.fallback_summary(selected_city, benchmarks)
            else:
                try:
                    structured = ai_summary.openai_summary(
                        api_key=api_key,
                        model_name=model_name,
                        selected_city=selected_city,
                        benchmarks=benchmarks,
                    )
                    structured = ai_summary.normalize_summary(structured, selected_city, benchmarks)
                except Exception:
                    structured = ai_summary.fallback_summary(selected_city, benchmarks)

            summary_text = ai_summary.combined_summary(structured)
            cache_payload = {
                "city": selected_city["city"],
                "state": selected_city["state"],
                "city_key": ai_summary.city_key(selected_city),
                "data_fingerprint": data_fingerprint,
                "model_name": model_name,
                "prompt_version": ai_summary.PROMPT_VERSION,
                "summary_text": summary_text,
                "strength_sentence": structured["strength_sentence"],
                "weakness_sentence": structured["weakness_sentence"],
                "comparison_sentence": structured["comparison_sentence"],
                "generated_at": generated_at,
                "cre_sentences_text": "",
                "research_sources_json": "[]",
                "cre_generated_at": None,
                "cre_research_version": None,
                "cre_result_type": None,
                "cre_failure_category": None,
                "cre_failure_code": None,
                "cre_failure_param": None,
            }

            try:
                db_module.upsert_city_summary_cache(conn, cache_payload)
            except Exception as exc:
                current_app.logger.warning(
                    "FIRE Metrics city-summary cache write failed: %s",
                    exc.__class__.__name__,
                )

            return jsonify({
                "status": "ready",
                "summary": summary_text,
                "summary_structured": structured,
                "research_sources": [],
                "generated_at": generated_at,
                "data_refreshed_at": metadata.get("last_refresh_at"),
                "cached": False,
                "city_key": cache_payload["city_key"],
                "relative_market_profile_score": benchmarks.get("relative_market_profile_score"),
                "relative_market_profile_percentile": benchmarks.get("relative_market_profile_percentile"),
                "tracked_city_relative_market_profile_average": benchmarks.get("tracked_city_relative_market_profile_average"),
                "recommendation_category": benchmarks.get("recommendation_category"),
                "score": benchmarks.get("selected_overall_score"),
                "computed_composite_score": benchmarks.get("selected_overall_score"),
                "tracked_city_average": benchmarks.get("tracked_city_average"),
                "tracked_city_composite_average": benchmarks.get("tracked_city_average"),
                "tracked_city_count": benchmarks.get("tracked_city_count"),
                "percentile": benchmarks.get("selected_percentile"),
                "source": "generated",
                "cre_status": "skipped",
                "cre_failure_category": None,
                "cre_failure_code": None,
                "cre_failure_param": None,
            })
    except Exception as exc:
        current_app.logger.exception("FIRE Metrics city-summary endpoint failed: %s", exc.__class__.__name__)
        response = _summary_unavailable_response(
            selected_city=None,
            benchmark_data=None,
            reason="Summary generation is currently unavailable.",
            data_refreshed_at=None,
        )
        response["error_code"] = "summary_endpoint_failed"
        return jsonify(response), 500


@fire_metrics_bp.route("/api/city-summary-cre", methods=["POST"])
@login_required
def city_summary_cre():
    payload = request.get_json(silent=True) or {}
    cre_generation_allowed = _cre_generation_allowed(payload)
    city_key = str(payload.get("city_key") or "").strip()
    city = str(payload.get("city") or "").strip()
    state = str(payload.get("state") or "").strip().upper()

    if not city_key and (not city or not state):
        return jsonify({
            "status": "error",
            "error_code": "invalid_city_identifier",
            "user_message": "City identifier is required.",
        }), 400

    try:
        with db_module.get_connection() as conn:
            selected_city = db_module.fetch_city_by_summary_identity(
                conn,
                city_key=city_key or None,
                city=city or None,
                state=state or None,
            )
            if not selected_city:
                return jsonify({
                    "status": "error",
                    "error_code": "city_not_found",
                    "user_message": "City not found in tracked FIRE Metrics data.",
                }), 404

            all_cities = db_module.fetch_all_included_cities(conn)
            metadata = db_module.get_metadata(conn)
            benchmarks = ai_summary.compute_benchmarks(selected_city, all_cities)
            model_name = _summary_model_name()
            fingerprint_input = ai_summary.fingerprint_payload(
                selected_city=selected_city,
                benchmarks=benchmarks,
                model_name=model_name,
                refresh_last_at=metadata.get("last_refresh_at"),
            )
            data_fingerprint = ai_summary.build_fingerprint(fingerprint_input)

            cre_payload = _cre_payload_for_city(
                conn,
                selected_city=selected_city,
                data_fingerprint=data_fingerprint,
                model_name=model_name,
                cre_generation_allowed=cre_generation_allowed,
            )

            return jsonify({
                "status": "ready",
                "city_key": ai_summary.city_key(selected_city),
                "cre_status": cre_payload.get("cre_status") or "skipped",
                "cre_summary": cre_payload.get("cre_summary") or "",
                "research_sources": cre_payload.get("research_sources") or [],
                "cre_failure_category": cre_payload.get("cre_failure_category"),
                "cre_failure_code": cre_payload.get("cre_failure_code"),
                "cre_failure_param": cre_payload.get("cre_failure_param"),
            })
    except Exception as exc:
        current_app.logger.exception("FIRE Metrics city-summary-cre endpoint failed: %s", exc.__class__.__name__)
        return jsonify({
            "status": "error",
            "error_code": "cre_endpoint_failed",
            "user_message": "Institutional research is currently unavailable.",
        }), 500


@fire_metrics_bp.route("/refresh-status")
@login_required
def refresh_status():
    return jsonify(_refresh_status())


@fire_metrics_bp.route("/upload-crime-workbook", methods=["POST"])
@login_required
def upload_crime_workbook():
    if not _is_admin_user():
        return _admin_forbidden_response(is_ajax=request.headers.get("X-Requested-With") == "XMLHttpRequest")

    def safe_crime_workbook_status() -> dict:
        try:
            return _crime_workbook_status()
        except Exception:
            return {"exists": False, "uploaded_at": None}

    def respond(success: bool, message: str, status_code: int = 200):
        return jsonify({
            "success": success,
            "message": message,
            "crime_workbook": safe_crime_workbook_status(),
        }), status_code

    try:
        file = request.files.get("crime_workbook")
        if file is None or not file.filename:
            return respond(False, "No file selected.", 400)

        if not file.filename.lower().endswith(".xlsx"):
            return respond(False, "File must be a .xlsx workbook.", 400)

        data = file.read()
        if not data:
            return respond(False, "File is empty. Upload the .xlsx workbook exactly as downloaded from the FBI.", 400)

        if len(data) > MAX_CRIME_WORKBOOK_BYTES:
            size_mb = len(data) / (1024 * 1024)
            return respond(
                False,
                f"File is too large ({size_mb:.1f} MB) -- the real FBI Table 8 workbook is "
                f"only a few MB. Check you selected the right file.",
                400,
            )

        validation_error = _validate_crime_workbook_bytes(data)
        if validation_error:
            return respond(False, validation_error, 400)

        # Uses the same resolver the crime pipeline uses. In production
        # this should be FBI_CRIME_WORKBOOK_PATH on the persistent /data
        # volume, so the file survives redeploys the same way the SQLite
        # DB now does via FIRE_METRICS_DB_PATH.
        target_path = _get_crime_workbook_path()
        target_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile("wb", dir=target_path.parent, delete=False) as tmp:
                tmp.write(data)
                tmp_path = Path(tmp.name)
            os.replace(tmp_path, target_path)
        finally:
            if tmp_path and tmp_path.exists():
                tmp_path.unlink()

        return respond(
            True,
            "Crime workbook uploaded. It will be picked up the next time you run "
            "\"Refresh All Data\".",
        )
    except Exception as exc:
        return respond(False, f"Unexpected error while uploading: {exc}", 500)


@fire_metrics_bp.route("/debug-refresh")
@login_required
def debug_refresh():
    if not _is_admin_user():
        abort(403)

    # TEMPORARY diagnostic route -- added specifically to inspect the raw
    # refresh_metadata table (including per-step results the normal status
    # payload doesn't surface) without needing direct Railway console/DB
    # access. Not linked from any page; remove once the climate-risk
    # never-populates investigation is resolved.
    with db_module.get_connection() as conn:
        metadata = db_module.get_metadata(conn)
    steps_raw = metadata.get("refresh_steps_json")
    try:
        parsed_steps = json.loads(steps_raw) if steps_raw else None
    except json.JSONDecodeError as exc:
        parsed_steps = f"<could not parse refresh_steps_json: {exc}>"
    return jsonify({
        "raw_metadata": metadata,
        "parsed_steps": parsed_steps,
    })


@fire_metrics_bp.route("/download-latest")
@login_required
def download_latest():
    if not _is_admin_user():
        abort(403)

    data = _export_workbook()
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="fire_metrics_city_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@fire_metrics_bp.route("/export/city-analytics.xlsx", methods=["POST"])
@login_required
def export_city_analytics_excel():
    payload = request.get_json(silent=True) or {}
    rows = _normalize_city_analytics_export_rows(payload.get("cities"))
    if not rows:
        return jsonify({
            "status": "error",
            "error_code": "empty_city_analytics",
            "user_message": "Add at least one city to City Analytics before exporting.",
        }), 400

    data = _build_city_analytics_xlsx(rows)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name=_city_analytics_export_filename("xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@fire_metrics_bp.route("/export/city-analytics.pdf", methods=["POST"])
@login_required
def export_city_analytics_pdf():
    payload = request.get_json(silent=True) or {}
    rows = _normalize_city_analytics_export_rows(payload.get("cities"))
    if not rows:
        return jsonify({
            "status": "error",
            "error_code": "empty_city_analytics",
            "user_message": "Add at least one city to City Analytics before exporting.",
        }), 400

    data = _build_city_analytics_pdf(rows)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name=_city_analytics_export_filename("pdf"),
        mimetype="application/pdf",
    )
